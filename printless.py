#Libs

from bs4 import BeautifulSoup as soup
import glob
import os
import pandas as pd
from openpyxl import load_workbook

#Resource finder

folder_path = os.getcwd()
html_files = glob.glob(os.path.join(folder_path, '*.html'))
lxmls = []
input_file = 'input.xlsx'

for file_path, e in zip(html_files,range(len(html_files))):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            html_content = file.read()

        lxmls.append(soup(html_content, 'lxml'))
    except Exception as e:
        with open('error.log', 'a') as log_file:
            log_file.write(f"Error processing file {file_path}: {e}\n")        

#Function definitions

def extract_test_steps(data):
    steps = []

    try:
        for element in range(len(data)):
            if len(data) == 3:
                mains = data.find_all('big', class_='Heading3', string='Main Part of Test Case')
            else:
                mains = data[element].find_all('big', class_='Heading3', string='Main Part of Test Case')

            for e in range(len(mains)):
                test_case = mains[e].find_previous('td', class_=['TestcaseHeadingPositiveResult', 'TestcaseHeadingNegativeResult']).get_text(strip=True)
                div = mains[e].find_next('div', class_='Indentation')
                table = div.find('table', class_='ResultTable')
                
                if table:
                    step_name = None
                    step_status = None
                    expected_output = None
                    step_output = None

                    for row_num, row in enumerate(table.find_all('tr')):
                        heading = row.find('big', class_='Heading4')
                        break_ = False
                        if heading:
                            text = heading.get_text(strip=True)
                            if 'None' in text:
                                text = text.strip().rsplit(': ', 1)
                                step_name = text[0]
                                step_output = ''
                                expected_output = step_output
                                step_status = ''
                            elif 'Expected' not in text:
                                text = text.rsplit(': ', 1)
                                step_name = text[0].strip()
                                step_status = text[1].strip()
                                expected_output = ''
                                if 'Failed' in step_status:
                                    step_output = row.find_next('td', class_="DefaultCell").get_text(strip=True)
                                    if 'MaskSymbolOp' in step_output:
                                        tab = row.find_next('table', class_='InfoTableExpand')
                                        step_output = tab.find_all('tr')[-1].find_all('td')[2].get_text(strip=True)
                                    if row.find_next('td', class_='DefaultCell', string='Test aborted due to BreakOnFail behavior.'):# and (step_name in row.find_next('td', class_='DefaultCell', string='Test aborted due to BreakOnFail behavior.').parent.parent.find_all('big', class_='Heading4')):
                                        for item in row.find_next('td', class_='DefaultCell', string='Test aborted due to BreakOnFail behavior.').parent.parent.find_all('big', class_='Heading4'):
                                            if step_name in item.get_text(strip=True):
                                                break_ = True
                                else:
                                    step_output = expected_output
                            else:
                                text = text.rsplit(':', 1)
                                text[0:1] = text[0].split('Expected:',2)
                                step_name = text[0].strip()
                                step_status = text[2].strip()
                                expected_output = text[1].strip()
                                if 'Failed' in step_status:
                                    step_output = row.find_next('td', class_="DefaultCell").get_text(strip=True)
                                    if 'MaskSymbolOp' in step_output:
                                        tab = row.find_next('table', class_='InfoTableExpand')
                                        step_output = tab.find_all('tr')[-1].find_all('td')[2].get_text(strip=True)
                                    if row.find_next('td', class_='DefaultCell', string='Test aborted due to BreakOnFail behavior.'):
                                        for item in row.find_next('td', class_='DefaultCell', string='Test aborted due to BreakOnFail behavior.').parent.parent.find_all('big', class_='Heading4'):
                                            if step_name in item.get_text(strip=True):
                                                break_ = True
                                else:
                                    step_output = expected_output

                            steps.append({
                                'Test Case': test_case,
                                'Step Name': step_name,
                                'Status': step_status,
                                'Expected Output': expected_output,
                                'Output': step_output,
                                'BreakOnFail': break_
                            })

    except Exception as e:
        with open('error.log', 'a') as log_file:
            log_file.write(f'Error extracting information: {e}\n')

    return steps

def excel(steps, excel_file_path, new_file_name = f'autogenerated_report.xlsx'):
    try:
        df = pd.DataFrame(steps)

        wb = load_workbook(excel_file_path)
        sheet = wb.active

        excel_test_case_column = 2
        excel_step_name_column = 6
        excel_output_column = 8
        excel_status_column = 9
        Failed = False
        excel_test_case = excel_step_name ='xd' #inital val
        log = []

        for n_row, row in enumerate(sheet.iter_rows(min_row=3)):
            if row[excel_test_case_column - 1].value != None:
                excel_test_case = row[excel_test_case_column - 1].value
                Failed = False
            excel_step_name = row[excel_step_name_column - 1].value
            if isinstance(excel_test_case, str) and isinstance(excel_step_name, str):
                matched = False
                for index, df_row in df.iterrows():    
                    if (excel_test_case in df_row['Test Case']) and (excel_step_name in df_row['Step Name']):
                        matched = True
                        if df_row['BreakOnFail'] or Failed:
                            row[excel_status_column - 1].value = 'FAILED'
                            row[excel_output_column - 1].value = 'Break On Fail'
                            df = df.drop(index)
                            Failed = True
                        elif (df_row['Status'] == 'Passed') and not (Failed):
                            row[excel_status_column - 1].value = 'PASSED'
                            row[excel_output_column - 1].value = row[excel_output_column - 2].value
                            df = df.drop(index)
                        elif ((df_row['Status'] == '') and (row[excel_output_column - 2].value == None) or ('Passed' in df_row['Test Case']))and not (Failed):
                            row[excel_status_column - 1].value = 'PASSED'
                            row[excel_output_column - 1].value = row[excel_output_column - 2].value
                            df = df.drop(index)
                        else:
                            row[excel_output_column - 1].value = df_row['Output']
                            row[excel_status_column - 1].value = 'FAILED'   
                            df = df.drop(index)

                    
                if not matched:
                    if sheet.cell(row=row[0].row-1, column=excel_output_column).value == 'Break On Fail':
                        row[excel_output_column - 1].value = 'Break On Fail'
                    row[excel_status_column - 1].value = 'FAILED'
                    log.append(f'Row:{row[0].row+1},Test Case:{excel_test_case}, Step Name: {excel_step_name}\n')
                

        log_file = 'unmatched.log'
        with open(log_file, 'w') as logz:
            logz.writelines(log)


    except Exception as e:
        with open('error.log', 'a') as log_file:
            log_file.write(f'Error updating excel: {e}\nExcel row: {row[0].row}')

    return wb.save(new_file_name), log_file

#Programme

excel(extract_test_steps(lxmls), input_file)

print('Output saved to file -> autogenerated_report.xlsx \nLogs saved to file -> unmatched.log')